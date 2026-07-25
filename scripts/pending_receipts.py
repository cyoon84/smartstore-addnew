#!/usr/bin/env python3
"""북키퍼 장부(COGS 탭) ↔ 출고일 정산(cogs.json) 대사 — **정산에 아직 안 넣은 영수증** 찾기.

왜: 영수증은 `receipt-plus-address-poller`(20분마다) 가 자동으로 북키퍼 장부에 기입한다.
    그런데 정산 세션은 그 사실을 몰라서 사용자가 매번 "영수증 넣었어" 라고 알려줘야 했다.
    이 스크립트가 **장부를 진실의 원천**으로 삼아 미반영분을 뽑아준다 (SessionStart 훅이 자동 실행).

상태 파일: output/settlement/_receipts_reflected.json
    { "<날짜>|<merchant>|<total>": {"batch": "2026-07-24", "note": "...", "at": "..."} }
    batch 가 "재고" / "제외" 면 정산 반영 대상 아님(사입재고 전환·취소분 등)으로 간주.

사용:
  python3 scripts/pending_receipts.py                 # 미반영 목록 (사람용)
  python3 scripts/pending_receipts.py --json          # 기계용
  python3 scripts/pending_receipts.py --quiet         # 미반영 있을 때만 출력 (훅용)
  python3 scripts/pending_receipts.py --mark "2026-07-23|Amazon.ca|31.62" --batch 2026-07-24 --note "홍관표 1개분"
  python3 scripts/pending_receipts.py --mark-all-before 2026-07-22 --batch seed   # 과거분 일괄 시드
"""
import argparse, datetime, json, os, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEDGER = "/Volumes/External/claude/profit-expense-tracker/장부.xlsx"
COGS_TAB = "물건산거 (COGS)"
STATE = os.path.join(ROOT, "output", "settlement", "_receipts_reflected.json")
SHIP_WEEKDAYS = (1, 4)  # 화/금


def ship_day(date_str):
    y, m, d = map(int, str(date_str)[:10].split("-"))
    dt = datetime.date(y, m, d)
    for i in range(7):
        c = dt + datetime.timedelta(days=i)
        if c.weekday() in SHIP_WEEKDAYS:
            return c.isoformat()
    return date_str


def load_state():
    return json.load(open(STATE, encoding="utf-8")) if os.path.exists(STATE) else {}


def save_state(st):
    os.makedirs(os.path.dirname(STATE), exist_ok=True)
    json.dump(st, open(STATE, "w", encoding="utf-8"), ensure_ascii=False, indent=2, sort_keys=True)


def norm_date(v):
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10] if v else ""


def ledger_rows():
    from openpyxl import load_workbook
    if not os.path.exists(LEDGER):
        return []
    ws = load_workbook(LEDGER, data_only=True)[COGS_TAB]
    hdr = [c.value for c in ws[1]]
    ix = {h: i + 1 for i, h in enumerate(hdr) if h}
    out = []
    for r in range(2, ws.max_row + 1):
        date = norm_date(ws.cell(r, ix["날짜"]).value)
        merch = ws.cell(r, ix["merchant"]).value
        total = ws.cell(r, ix["total"]).value
        if not date or not merch:
            continue
        out.append({
            "key": f"{date}|{merch}|{total}",
            "date": date, "merchant": merch, "total": total,
            "method": ws.cell(r, ix.get("method of payment", 0)).value if ix.get("method of payment") else None,
            "receipt": ws.cell(r, ix.get("영수증파일", 0)).value if ix.get("영수증파일") else None,
            "note": ws.cell(r, ix.get("Note", 0)).value if ix.get("Note") else None,
            "batch_guess": ship_day(date),
        })
    return out


def pending(rows, st):
    return [r for r in rows if r["key"] not in st]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--quiet", action="store_true", help="미반영 있을 때만 출력(훅용)")
    ap.add_argument("--mark", action="append", default=[], metavar="KEY")
    ap.add_argument("--batch", help="반영한 출고일(YYYY-MM-DD) 또는 재고/제외/seed")
    ap.add_argument("--note", default="")
    ap.add_argument("--mark-all-before", metavar="YYYY-MM-DD", help="이 날짜 이전 장부행 일괄 시드")
    a = ap.parse_args()

    rows = ledger_rows()
    st = load_state()

    if a.mark or a.mark_all_before:
        if not a.batch:
            sys.exit("--batch 필요 (출고일 / 재고 / 제외 / seed)")
        keys = set(a.mark)
        if a.mark_all_before:
            keys |= {r["key"] for r in rows if r["date"] < a.mark_all_before}
        now = datetime.datetime.now().isoformat(timespec="seconds")
        for k in keys:
            st[k] = {"batch": a.batch, "note": a.note, "at": now}
        save_state(st)
        print(f"[mark] {len(keys)}건 → batch={a.batch}")
        return

    pend = pending(rows, st)
    if a.json:
        print(json.dumps(pend, ensure_ascii=False, indent=2))
        return
    if not pend:
        if not a.quiet:
            print("미반영 영수증 없음 (장부 COGS 전부 정산 반영/처리 완료)")
        return
    print(f"⚠️ 정산 미반영 영수증 {len(pend)}건 — 북키퍼 장부(COGS)에는 있는데 정산에 안 들어감")
    for r in pend:
        print(f"  · {r['date']} | {r['merchant']} | ${r['total']} → 출고일 후보 {r['batch_guess']}")
        if r["note"]:
            print(f"      {r['note']}")
        print(f"      key: {r['key']}")
    print("  처리: order_settlement.py --add-cogs 로 반영 후 "
          "pending_receipts.py --mark \"<key>\" --batch <출고일> (재고전환이면 --batch 재고)")


if __name__ == "__main__":
    main()
