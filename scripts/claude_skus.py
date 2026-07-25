#!/usr/bin/env python3
"""클로드 등록분(리스팅) 매출 추적 — 상품번호 레지스트리.

왜: product_info 에 live_product_id 가 거의 안 박혀서, 정산 때 "이 주문이 클로드가 등록한
    제품인가"를 매번 제목 매칭으로 헤맸다(부정확). → **상품번호를 한 번 분류(claude/prior)해
    두면 그 다음 주문부터 영구 자동 집계**. (pending_receipts.py 와 같은 대사+마킹 패턴.)

레지스트리: output/settlement/_sku_registry.json
    { "<상품번호>": {"origin":"claude"|"prior", "slug":"", "name":"", "at":""} }
    origin=claude → 클로드가 리스팅(상품명·상세·태그) 작성한 제품.
    origin=prior  → 클로드 도입 이전 등록분(또는 사용자 수동 등록).

사용:
  python3 scripts/claude_skus.py --classify <발주발송.xlsx>   # 미분류 상품번호 + new-item 매칭 제안
  python3 scripts/claude_skus.py --tag 12345678 --claude --slug downy_..._680g
  python3 scripts/claude_skus.py --tag 87654321 --prior --name "거버 이유식"
  python3 scripts/claude_skus.py --list [--origin claude]
  # order_settlement.py 가 이 레지스트리를 읽어 각 정산 md 에 "클로드 등록분 매출"을 자동 표시.
"""
import argparse, glob, io, json, os, re, datetime, unicodedata

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REG = os.path.join(ROOT, "output", "settlement", "_sku_registry.json")
NEWITEM = os.path.join(ROOT, "output", "new-item")


def load():
    return json.load(open(REG, encoding="utf-8")) if os.path.exists(REG) else {}


def save(r):
    os.makedirs(os.path.dirname(REG), exist_ok=True)
    json.dump(r, open(REG, "w", encoding="utf-8"), ensure_ascii=False, indent=2, sort_keys=True)


def classify(pid, reg):
    """상품번호 → 'claude' | 'prior' | None(미분류)."""
    e = reg.get(str(pid))
    return e["origin"] if e else None


# --- new-item 한글 제목 인덱스 (분류 제안용) ---
def newitem_index():
    idx = []
    for f in glob.glob(os.path.join(NEWITEM, "*", "*product_info.json")):
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        slug = os.path.basename(os.path.dirname(f))
        titles = set()

        def dig(o):
            if isinstance(o, dict):
                for k, v in o.items():
                    if k in ("title_ko", "product_name_ko", "group_name_ko") and isinstance(v, str):
                        titles.add(v)
                    dig(v)
            elif isinstance(o, list):
                for x in o:
                    dig(x)
        dig(d)
        idx.append((slug, titles))
    return idx


def _toks(s):
    return set(re.findall(r"[가-힣A-Za-z0-9]{2,}", unicodedata.normalize("NFC", str(s or ""))))


def suggest(name, idx):
    a = _toks(name)
    best = None
    for slug, titles in idx:
        for t in titles:
            c = a & _toks(t)
            if len(c) >= 3 and (best is None or len(c) > best[1]):
                best = (slug, len(c))
    return best[0] if best else None


def read_orders(path, pw="1111"):
    import msoffcrypto
    from openpyxl import load_workbook
    buf = io.BytesIO()
    with open(path, "rb") as fh:
        try:
            o = msoffcrypto.OfficeFile(fh); o.load_key(password=pw); o.decrypt(buf)
        except Exception:
            fh.seek(0); buf = io.BytesIO(fh.read())
    buf.seek(0)
    ws = load_workbook(buf, data_only=True).active
    rows = [r for r in ws.iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows) if r and any(str(c).strip() == "상품주문번호" for c in r if c)), 0)
    h = [str(c).strip() if c is not None else "" for c in rows[hi]]
    def ci(*n):
        for x in n:
            if x in h: return h.index(x)
    ip, inm, ibd = ci("상품번호(스마트스토어)", "상품번호"), ci("상품명"), ci("정산예정금액")
    out = {}
    for r in rows[hi + 1:]:
        if not r or not r[0]: continue
        pid = str(r[ip]) if ip is not None and r[ip] else ""
        if not pid: continue
        d = out.setdefault(pid, {"name": r[inm], "bd": 0, "n": 0})
        d["bd"] += r[ibd] or 0; d["n"] += 1
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--classify", metavar="발주발송.xlsx")
    ap.add_argument("--password", default="1111")
    ap.add_argument("--tag", metavar="상품번호")
    ap.add_argument("--claude", action="store_true")
    ap.add_argument("--prior", action="store_true")
    ap.add_argument("--slug", default="")
    ap.add_argument("--name", default="")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--origin", choices=["claude", "prior"])
    a = ap.parse_args()
    reg = load()

    if a.tag:
        if not (a.claude or a.prior):
            raise SystemExit("--claude 또는 --prior 지정")
        reg[str(a.tag)] = {"origin": "claude" if a.claude else "prior",
                           "slug": a.slug, "name": a.name,
                           "at": datetime.datetime.now().isoformat(timespec="seconds")}
        save(reg)
        print(f"[tag] {a.tag} → {reg[str(a.tag)]['origin']}" + (f" ({a.slug or a.name})" if (a.slug or a.name) else ""))
        return

    if a.list:
        for pid, e in sorted(reg.items()):
            if a.origin and e["origin"] != a.origin: continue
            print(f"  {pid}  {e['origin']:<6}  {e.get('slug') or e.get('name','')}")
        print(f"— 총 {sum(1 for e in reg.values() if not a.origin or e['origin']==a.origin)}건")
        return

    if a.classify:
        orders = read_orders(a.classify, a.password)
        idx = newitem_index()
        unclassified = [(pid, d) for pid, d in orders.items() if classify(pid, reg) is None]
        if not unclassified:
            print("미분류 상품번호 없음 (이 배치 전부 분류 완료)")
            return
        print(f"⚠️ 미분류 상품번호 {len(unclassified)}건 — claude/prior 로 분류하세요:")
        for pid, d in sorted(unclassified, key=lambda x: -x[1]["bd"]):
            sg = suggest(d["name"], idx)
            tip = f"  → 매칭 제안: {sg} (claude 가능성)" if sg else "  → new-item 매칭 없음 (prior 가능성)"
            print(f"  {pid} | ₩{d['bd']:>8,.0f} ({d['n']}) | {str(d['name'])[:40]}{tip}")
            if sg:
                print(f"      python3 scripts/claude_skus.py --tag {pid} --claude --slug {sg}")
            else:
                print(f"      python3 scripts/claude_skus.py --tag {pid} --prior --name \"{str(d['name'])[:30]}\"")
        return

    ap.print_help()


if __name__ == "__main__":
    main()
