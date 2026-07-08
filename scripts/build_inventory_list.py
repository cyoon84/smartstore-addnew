#!/usr/bin/env python3
"""미리 사입(재고 보유)한 제품의 **재고 관리 엑셀**을 생성한다.

등록 제품 폴더(output/new-item/<slug>/<slug>_product_info.json)에서
제품명·매입처·원가·판매가·마진·카테고리·판매URL 등 **자동 채움 값**을 뽑고,
사입일·사입수량·판매수량·현재고(수식)·유통기한·보관위치·재입고·메모 등
**손으로 채울 칸**을 비워 둔 관리 시트를 만든다.

인자는 슬러그(폴더명) 또는 한글 제품명 일부 둘 다 받는다(부분일치).

사용:
  python3 scripts/build_inventory_list.py <slug 또는 이름> [<slug 또는 이름> ...]
  python3 scripts/build_inventory_list.py downy_unstopables_paradise_853g twizzlers
  python3 scripts/build_inventory_list.py --out output/inventory/재고관리_2026-07-07.xlsx 다우니 트위즐러
  python3 scripts/build_inventory_list.py --list-file 사입목록.txt      # 한 줄당 slug/이름
  python3 scripts/build_inventory_list.py --all                        # 등록 제품 전체(마스터)

기존 재고 엑셀에 이어붙이려면 --append <기존.xlsx> — 이미 있는 슬러그는 건너뛰고 신규만 추가.
"""
import argparse, csv, glob, json, os, sys, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
NEW_ITEM = os.path.join(ROOT, "output", "new-item")
GUIDE_DIR = os.path.join(ROOT, "guide")
STORE_URL = "https://smartstore.naver.com/finchmart_ca/products/{}"

_GUIDE_CACHE = None


def _sales_csvs(guide_dir):
    """guide 판매목록 CSV — Product_*.csv + 스마트스토어상품_*.csv 둘 다, 파일명 날짜(YYYYMMDD_HHMMSS)순(오래된→최신). macOS 한글파일명 NFD/NFC 이슈 회피 위해 *.csv 글롭 후 NFC 정규화로 필터. (네이버 내보내기 이름 변경 2026-07-08 대응.)"""
    import re as _re, glob as _glob, os as _os, unicodedata as _ud
    out = []
    for p in _glob.glob(_os.path.join(guide_dir, "*.csv")):
        b = _ud.normalize("NFC", _os.path.basename(p))
        if b.startswith("Product_") or b.startswith("스마트스토어상품_"):
            out.append(p)
    def _k(p):
        m = _re.search(r"\d{8}_\d{6}", _os.path.basename(p))
        return m.group(0) if m else _os.path.basename(p)
    return sorted(out, key=_k)


def guide_index():
    """guide/Product_*.csv (전체 판매 제품 리스트)를 병합해 상품번호→행 dict.
    같은 상품이 여러 CSV에 있으면 가장 최근 파일 값을 채택."""
    global _GUIDE_CACHE
    if _GUIDE_CACHE is not None:
        return _GUIDE_CACHE
    by_id = {}
    for f in _sales_csvs(GUIDE_DIR):  # 오래된→최신 (Product_ + 스마트스토어상품_)
        try:
            with open(f, encoding="utf-8-sig", errors="replace") as fh:
                rows = list(csv.reader(fh))
        except Exception:
            continue
        if not rows:
            continue
        header = rows[0]
        idx = {h: i for i, h in enumerate(header)}
        def g(row, h):
            return row[idx[h]] if h in idx and idx[h] < len(row) else ""
        for row in rows[1:]:
            pid = g(row, "상품번호(스마트스토어)").strip()
            if not pid:
                continue
            by_id[pid] = {                        # 최신 파일이 마지막에 덮어씀
                "id": pid,
                "name": g(row, "상품명").strip(),
                "price": g(row, "판매가").strip(),
                "stock": g(row, "재고수량").strip(),
                "status": g(row, "판매상태").strip(),
                "group": g(row, "그룹상품번호").strip(),
                "brand": g(row, "브랜드명").strip(),
                "maker": g(row, "제조사명").strip(),
            }
    _GUIDE_CACHE = by_id
    return by_id


def row_from_guide(gd):
    """guide CSV 항목 → 재고 시트 행(이름·판매가·판매URL·스토어재고). 원가·매입처·마진은 수동."""
    price = gd.get("price") or ""
    try:
        price = int(float(price))
    except (ValueError, TypeError):
        pass
    memo = []
    if gd.get("status") and gd["status"] != "판매중":
        memo.append(gd["status"])   # 판매종료·품절 등만 표시 (스토어 노출재고는 실재고와 무관 → 제외)
    return {
        "상품명": gd.get("name", ""),
        "판매가(₩)": price,
        "판매URL": STORE_URL.format(gd["id"]) if gd.get("id") else "",
        "슬러그": gd.get("id", ""),
        "메모": " · ".join(memo),
    }

# 재고관리 시트 컬럼 정의: (헤더, 폭, 자동채움여부)
COLUMNS = [
    ("사입일",        12, False),   # A  YYYY-MM-DD
    ("상품명",        40, True),    # B
    ("매입처",        14, True),    # C  코스트코/월마트 등
    ("용량/규격",     16, True),    # D
    ("원가(현지)",    12, True),    # E  CAD 등
    ("원가통화",      8,  True),    # F
    ("판매가(₩)",     11, True),    # G
    ("마진",          14, True),    # H
    ("사입단가(₩)",   12, False),   # I  실제 매입 단가(수동)
    ("사입수량",      9,  False),   # J
    ("판매수량",      9,  False),   # K
    ("현재고",        9,  False),   # L  = 사입수량 - 판매수량 (수식)
    ("유통기한",      12, False),   # M
    ("보관위치",      12, False),   # N
    ("재입고필요",    10, False),   # O
    ("판매URL",       22, True),    # P
    ("카테고리",      30, True),    # Q
    ("슬러그",        30, True),    # R  (키/추적용)
    ("메모",          30, False),   # S
]


def pick(d, *keys):
    for k in keys:
        if k is None:
            continue
        v = d.get(k) if isinstance(k, str) else k
        if v not in (None, "", [], {}):
            return v
    return ""


def extract(pinfo, slug):
    """product_info.json → 재고 시트 행 dict (자동 채움 값만)."""
    pricing = pinfo.get("pricing") or {}
    calc = pinfo.get("calculation") or {}
    brand_ko = pinfo.get("brand_ko") or pinfo.get("brand") or ""

    title = pick(pinfo, "product_name_ko", "title_ko", "group_name_ko")
    source = pick(pinfo, "source_store")
    weight = pick(pinfo, "net_weight", "variant")

    cost = pick(pricing, "cost_original", "cost_with_tax") or pick(pinfo, "cost_original")
    cost_cur = pick(pricing, "cost_currency") or ("CAD" if cost != "" else "")

    sell = (pick(pricing, "sell_price_krw", "sell_krw")
            or pick(pinfo, "sell_price_krw", "selling_price_krw", "sell_krw")
            or pick(calc, "sell_price_krw", "sell_krw"))

    # 마진: markup(±통화) 또는 verification.realized_margin_cad
    ver = pricing.get("verification") or {}
    markup = pick(pricing, "markup")
    mcur = pick(pricing, "markup_currency") or (cost_cur if markup != "" else "")
    if markup != "":
        margin = f"{markup} {mcur}".strip()
    elif ver.get("realized_margin_cad") not in (None, ""):
        margin = f"{ver['realized_margin_cad']} CAD"
    else:
        margin = ""

    url = pick(pinfo, "live_url")
    cat = pick(pinfo, "category_proposed", "naver_category", "category_naver",
               "category_naver_confirmed")

    return {
        "상품명": title,
        "매입처": source,
        "용량/규격": weight,
        "원가(현지)": cost,
        "원가통화": cost_cur,
        "판매가(₩)": sell,
        "마진": margin,
        "판매URL": url,
        "카테고리": cat,
        "슬러그": slug,
    }


def all_slugs():
    out = []
    for fo in sorted(glob.glob(os.path.join(NEW_ITEM, "*", ""))):
        name = os.path.basename(os.path.dirname(fo))
        if name.startswith("_"):
            continue
        if glob.glob(os.path.join(fo, "*product_info.json")):
            out.append(name)
    return out


def resolve_targets(tokens):
    """slug 정확일치 → slug/제품명 부분일치. 반환: (resolved_pairs, manual_names).
    - 정확/단일일치: (slug, product_info) 로 자동채움.
    - 후보 여러 개(모호): 후보 출력하고 건너뜀(사용자가 좁히게).
    - 매칭 0개: 미등록 제품으로 보고 상품명만 채운 **수동 행**으로 추가(manual_names)."""
    slugs = all_slugs()
    # 제품명 인덱스 (부분일치용)
    name_index = {}
    for s in slugs:
        pj = glob.glob(os.path.join(NEW_ITEM, s, "*product_info.json"))
        try:
            d = json.load(open(pj[0]))
        except Exception:
            d = {}
        name_index[s] = (pick(d, "product_name_ko", "title_ko", "group_name_ko") or "", d)

    resolved, guide_rows, manual = [], [], []
    seen = set()
    for tok in tokens:
        t = tok.strip()
        if not t:
            continue
        hit = None
        if t in name_index:                       # slug 정확일치
            hit = t
        else:
            cands = [s for s in slugs
                     if t.lower() in s.lower() or t in name_index[s][0]]
            if len(cands) == 1:
                hit = cands[0]
            elif len(cands) > 1:
                print(f"[?] '{t}' 후보 여러 개 — 슬러그로 좁혀주세요:")
                for c in cands[:12]:
                    print(f"      {c}  ({name_index[c][0]})")
                continue
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append((hit, name_index[hit][1]))
            continue
        if hit:
            continue
        # output/new-item 실패 → guide 판매목록(전체) 조회 (상품ID 숫자 또는 이름 부분일치)
        gidx = guide_index()
        gd = None
        if t.isdigit() and t in gidx:
            gd = gidx[t]
        else:
            gcands = [v for v in gidx.values() if t in v["name"]]
            if len(gcands) == 1:
                gd = gcands[0]
            elif len(gcands) > 1:
                print(f"[?] '{t}' guide 후보 여러 개 — 상품ID로 좁혀주세요:")
                for c in gcands[:12]:
                    print(f"      {c['id']}  ({c['name']})")
                continue
        if gd and gd["id"] not in seen:
            seen.add(gd["id"])
            guide_rows.append(row_from_guide(gd))
        elif not gd:
            manual.append(t)                      # 미등록 → 수동 행
    if guide_rows:
        print(f"[i] guide 판매목록에서 채움: {len(guide_rows)}개")
    if manual:
        print(f"[i] 미등록(수동 행) 추가: {', '.join(manual)}")
    return resolved, guide_rows, manual


def build_workbook(rows, out_path, existing_slugs=None):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    existing_slugs = existing_slugs or set()
    headers = [c[0] for c in COLUMNS]

    if os.path.exists(out_path) and existing_slugs:
        wb = load_workbook(out_path)
        ws = wb.active
        start = ws.max_row + 1
        new_only = [r for r in rows if r.get("슬러그") not in existing_slugs]
        rows = new_only
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "재고관리"
        # 헤더
        head_fill = PatternFill("solid", fgColor="E0483F")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="F0C9C4")
        for ci, (h, w, _auto) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin, right=thin)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
        ws.row_dimensions[1].height = 30
        start = 2

    from openpyxl.utils import get_column_letter as gcl
    col_idx = {h: i + 1 for i, (h, _, _) in enumerate(COLUMNS)}
    cur_c = gcl(col_idx["현재고"])
    buy_c = gcl(col_idx["사입수량"])
    sold_c = gcl(col_idx["판매수량"])

    for i, r in enumerate(rows):
        rownum = start + i
        for h, ci in col_idx.items():
            ws.cell(row=rownum, column=ci, value=r.get(h, ""))
        # 현재고 = 사입수량 - 판매수량 (둘 다 채워졌을 때만 표시)
        ws.cell(row=rownum, column=col_idx["현재고"],
                value=f'=IF(ISNUMBER({buy_c}{rownum}),{buy_c}{rownum}-IF(ISNUMBER({sold_c}{rownum}),{sold_c}{rownum},0),"")')

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("targets", nargs="*", help="slug 또는 한글 제품명 일부")
    ap.add_argument("--out", help="출력 xlsx 경로")
    ap.add_argument("--list-file", help="한 줄당 slug/이름이 든 텍스트 파일")
    ap.add_argument("--all", action="store_true", help="등록 제품 전체(마스터)")
    ap.add_argument("--append", help="기존 재고 엑셀에 신규 슬러그만 이어붙임")
    args = ap.parse_args()

    tokens = list(args.targets)
    if args.list_file:
        with open(args.list_file, encoding="utf-8") as f:
            tokens += [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]

    if args.all:
        pairs = []
        for s in all_slugs():
            pj = glob.glob(os.path.join(NEW_ITEM, s, "*product_info.json"))
            try:
                pairs.append((s, json.load(open(pj[0]))))
            except Exception:
                pass
        guide_rows, manual = [], []
    else:
        if not tokens:
            sys.exit("사입한 제품 slug/이름을 인자로 주거나 --list-file / --all 사용")
        pairs, guide_rows, manual = resolve_targets(tokens)

    if not pairs and not guide_rows and not manual:
        sys.exit("[!] 매칭된 제품이 없습니다.")

    rows = [extract(d, s) for s, d in pairs]        # 등록 산출물: 원가·마진까지 자동
    rows += guide_rows                              # guide 판매목록: 이름·판매가·URL 자동
    rows += [{"상품명": name} for name in manual]   # 미등록: 상품명만

    out = args.append or args.out or os.path.join(
        ROOT, "output", "inventory",
        f"재고관리_{datetime.date.today().isoformat()}.xlsx")

    existing = set()
    if args.append and os.path.exists(args.append):
        from openpyxl import load_workbook
        ws = load_workbook(args.append).active
        scol = [c[0] for c in COLUMNS].index("슬러그") + 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=scol).value
            if v:
                existing.add(str(v))

    n = build_workbook(rows, out, existing_slugs=existing)
    print(f"[✓] 재고관리 엑셀: {out}")
    print(f"    {n}개 행 {'추가' if existing else '생성'} · 자동채움: 제품명·매입처·원가·판매가·마진·카테고리·URL")
    if existing:
        print(f"    (기존 {len(existing)}개 슬러그는 건너뜀)")
    print("    손입력: 사입일·사입단가·사입수량·판매수량·유통기한·보관위치·재입고·메모 (현재고는 수식 자동)")


if __name__ == "__main__":
    main()
