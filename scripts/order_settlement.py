#!/usr/bin/env python3
"""네이버 스마트스토어 **발주발송관리** 엑셀(암호 걸림) → 출고일 정산·손익 계산 + 재고 판매반영.

매 출고일(화/금) 워크플로:
  1) 발주발송 엑셀에서 **매출(정산 예정)** 자동 계산
       매출 = Σ정산예정금액(BD, 제품 정산=상품가−Npay수수료−매출연동수수료)
              + Σ배송비(묶음번호로 유니크) × (1−0.0363)   # 배송비엔 네이버 수수료 3.63%만 (§7-2)
  2) (선택) 재고관리 엑셀에 판매수량 반영 (--update-inventory)  ⚠️ 출고파일당 1회만
  3) **물건값(COGS)** 은 영수증 올 때마다 구성요소로 누적 (--add-cogs "라벨=금액"),
     **실배송비** 는 한미 리포트 오면 (--shipping-cost) → **순이익 = 매출 − 물건값 − 실배송비**
  결과: output/settlement/정산_<날짜>.md (+ 물건값 구성요소 정산_<날짜>_cogs.json)

사용:
  python3 scripts/order_settlement.py <발주발송.xlsx> [--password 1111]
  python3 scripts/order_settlement.py <파일> --update-inventory output/inventory/재고관리_2026-07-07.xlsx
  python3 scripts/order_settlement.py <파일> --add-cogs "애프터바이트(박은영,재고)=16070"   # 물건값 요소 추가
  python3 scripts/order_settlement.py <파일> --shipping-cost 45000     # 한미 실배송비 → 순이익
  # --date 로 출고일 지정(기본: 파일명 날짜). --cogs 로 구성요소 무시하고 총액 직접 지정.
"""
import argparse, io, json, os, sys, datetime
from collections import defaultdict, OrderedDict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SETTLE_DIR = os.path.join(ROOT, "output", "settlement")
SHIP_FEE_RATE = 0.0363   # 배송비 네이버 수수료(Npay만) — 제품가 6.6%와 다름 (§7-2)


def load_orders(path, password):
    import msoffcrypto
    from openpyxl import load_workbook
    buf = io.BytesIO()
    with open(path, "rb") as fh:
        try:
            off = msoffcrypto.OfficeFile(fh)
            off.load_key(password=password)
            off.decrypt(buf)
        except Exception:
            fh.seek(0); buf = io.BytesIO(fh.read())
    buf.seek(0)
    ws = load_workbook(buf, data_only=True).active
    rows = [r for r in ws.iter_rows(values_only=True)]
    hidx = next((i for i, r in enumerate(rows)
                 if r and any(str(c).strip() == "상품주문번호" for c in r if c)), 0)
    header = [str(c).strip() if c is not None else "" for c in rows[hidx]]
    data = [r for r in rows[hidx + 1:] if r and r[0]]
    return header, data


def col(header, *names):
    for n in names:
        for i, h in enumerate(header):
            if h == n:
                return i
    for n in names:
        for i, h in enumerate(header):
            if n in h:
                return i
    return None


def compute(header, data):
    c_recip = col(header, "수취인명")
    c_pno   = col(header, "상품번호(스마트스토어)", "상품번호")
    c_name  = col(header, "상품명")
    c_qty   = col(header, "수량")
    c_sgrp  = col(header, "배송비 묶음번호")
    c_sfee  = col(header, "배송비 합계")
    c_bd    = col(header, "정산예정금액")
    bd_sum = 0
    per_recip = defaultdict(lambda: {"bd": 0, "ship": 0})
    ship_by_group = OrderedDict()
    items = []
    for r in data:
        bd = r[c_bd] or 0
        bd_sum += bd
        recip = r[c_recip]
        per_recip[recip]["bd"] += bd
        g = r[c_sgrp]
        if g not in ship_by_group:
            ship_by_group[g] = (r[c_sfee] or 0, recip)
        items.append({"recip": recip, "pno": str(r[c_pno]) if r[c_pno] else "",
                      "name": r[c_name], "qty": r[c_qty] or 1})
    for g, (fee, recip) in ship_by_group.items():
        per_recip[recip]["ship"] += fee
    ship_total = sum(f for f, _ in ship_by_group.values())
    ship_settle = round(ship_total * (1 - SHIP_FEE_RATE))
    return {"bd_sum": bd_sum, "ship_total": ship_total, "ship_settle": ship_settle,
            "revenue": bd_sum + ship_settle, "per_recip": per_recip,
            "items": items, "ship_by_group": ship_by_group}


def update_inventory(inv_path, items):
    from openpyxl import load_workbook
    if not os.path.exists(inv_path):
        return "재고파일 없음"
    wb = load_workbook(inv_path); ws = wb.active
    hdr = [c.value for c in ws[1]]; cix = {h: i + 1 for i, h in enumerate(hdr)}
    scol = cix["슬러그"]
    row_of = {str(ws.cell(rr, scol).value): rr for rr in range(2, ws.max_row + 1)
              if ws.cell(rr, scol).value is not None}
    updated = []
    for it in items:
        rr = row_of.get(it["pno"])
        if rr:
            cur = ws.cell(rr, cix["판매수량"]).value or 0
            ws.cell(rr, cix["판매수량"]).value = int(cur) + int(it["qty"])
            updated.append((it["name"], it["qty"]))
    if updated:
        wb.save(inv_path)
    return updated


def cogs_file(date):
    return os.path.join(SETTLE_DIR, f"정산_{date}_cogs.json")


def load_cogs(date):
    p = cogs_file(date)
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else []


def save_cogs(date, comps):
    os.makedirs(SETTLE_DIR, exist_ok=True)
    json.dump(comps, open(cogs_file(date), "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def render_md(date, s, cogs_comps, cogs_total, ship_cost):
    L = [f"# 출고일 정산 — {date}", ""]
    L.append(f"## 매출 (정산 예정) = **₩{s['revenue']:,}**")
    L.append(f"- 제품 정산예정(ΣBD): ₩{s['bd_sum']:,}")
    L.append(f"- 배송비 정산: ₩{s['ship_settle']:,} (배송비 ₩{s['ship_total']:,} × (1−3.63%), 묶음 유니크)")
    L += ["", "### 수취인별", "| 수취인 | 제품정산 | 배송정산 | 소계 |", "|---|---|---|---|"]
    for rc, v in s["per_recip"].items():
        sh_a = round(v["ship"] * (1 - SHIP_FEE_RATE))
        L.append(f"| {rc} | ₩{v['bd']:,} | ₩{sh_a:,} (배송₩{v['ship']:,}) | ₩{v['bd']+sh_a:,} |")
    L += ["", "## 물건값 (COGS)"]
    if cogs_comps:
        for c in cogs_comps:
            L.append(f"- {c['label']}: ₩{c['amount']:,}")
        L.append(f"- **소계: ₩{cogs_total:,}**" + ("" if not any('대기' in c['label'] for c in cogs_comps) else " (일부 대기)"))
    elif cogs_total is not None:
        L.append(f"- 총액(직접입력): ₩{cogs_total:,}")
    else:
        L.append("- (영수증 대기)")
    L += ["", "## 손익"]
    L.append(f"- 매출(정산): ₩{s['revenue']:,}")
    L.append(f"- − 물건값(COGS): {('₩'+format(cogs_total, ',')) if cogs_total is not None else '(영수증 대기)'}")
    L.append(f"- − 실 배송비(한미): {('₩'+format(ship_cost, ',')) if ship_cost is not None else '(한미 리포트 대기)'}")
    if cogs_total is not None and ship_cost is not None:
        L.append(f"- **= 순이익: ₩{s['revenue']-cogs_total-ship_cost:,}**")
    else:
        L.append("- **= 순이익: (물건값·실배송비 입력 후 확정)**")
    L += ["", f"_주문 {len(s['items'])}건 · 배송 묶음 {len(s['ship_by_group'])}개 · 갱신 {date}_"]
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("orders_xlsx")
    ap.add_argument("--password", default="1111")
    ap.add_argument("--update-inventory", metavar="재고.xlsx")
    ap.add_argument("--add-cogs", action="append", default=[], metavar="라벨=금액",
                    help="물건값 구성요소 추가(누적). 반복 가능.")
    ap.add_argument("--cogs", type=int, help="물건값 총액 직접지정(구성요소 무시)")
    ap.add_argument("--shipping-cost", type=int, help="총 실 배송비(원, 한미)")
    ap.add_argument("--date")
    args = ap.parse_args()

    header, data = load_orders(args.orders_xlsx, args.password)
    s = compute(header, data)

    date = args.date
    if not date:
        import re
        m = re.search(r"(\d{8})", os.path.basename(args.orders_xlsx))
        date = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:8]}" if m else "unknown"

    # 물건값 구성요소 누적
    comps = load_cogs(date)
    for spec in args.add_cogs:
        label, _, amt = spec.rpartition("=")
        comps.append({"label": label.strip(), "amount": int(amt.strip())})
    if args.add_cogs:
        save_cogs(date, comps)

    if args.cogs is not None:
        cogs_total = args.cogs
    elif comps:
        cogs_total = sum(c["amount"] for c in comps)
    else:
        cogs_total = None

    print(f"[정산] {date} · 주문 {len(s['items'])}건 · 배송묶음 {len(s['ship_by_group'])}개")
    print(f"  매출(정산 예정) = ΣBD {s['bd_sum']:,} + 배송정산 {s['ship_settle']:,} = ₩{s['revenue']:,}")
    if cogs_total is not None:
        print(f"  물건값(COGS) = ₩{cogs_total:,}" + (f" ({len(comps)}개 요소)" if comps else ""))
    if cogs_total is not None and args.shipping_cost is not None:
        print(f"  순이익 = {s['revenue']:,} − {cogs_total:,} − {args.shipping_cost:,} = ₩{s['revenue']-cogs_total-args.shipping_cost:,}")

    if args.update_inventory:
        up = update_inventory(args.update_inventory, s["items"])
        if isinstance(up, list):
            print(f"  재고 반영: {len(up)}건 " + (", ".join(f"{n}×{q}" for n, q in up) if up else "(매칭 없음)"))
        else:
            print(f"  재고 반영 스킵: {up}")

    os.makedirs(SETTLE_DIR, exist_ok=True)
    out = os.path.join(SETTLE_DIR, f"정산_{date}.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write(render_md(date, s, comps, cogs_total, args.shipping_cost))
    print(f"  → {out}")


if __name__ == "__main__":
    main()
