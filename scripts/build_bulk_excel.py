#!/usr/bin/env python3
"""제품 산출물(product_info.json + detail.html)을 네이버 스마트스토어 **일괄등록 엑셀** 한 줄로 변환한다.

가격계산·저장처럼 메인 오케스트레이터가 소유하는 결정론적 변환 단계.
listing-writer 가 만든 콘텐츠(상품명·detail.html) + 확정 가격을 받아 템플릿 칸에 매핑한다.

- 템플릿:        guide/일괄등록 guide/ExcelSaveTemplate_*.xlsx  (가장 최근 파일)
- 카테고리코드:  guide/일괄등록 guide/category_*.xls            (카테고리 경로 → 코드 자동 해석)
- 원산지코드:    기본 캐나다(0204006) 고정 — 전 상품 캐나다로 깔고 사용자가 수동 보정. bulk.origin_code 로만 덮어씀.
- 택배사코드:    guide/일괄등록 guide/delivery-companies_*.xls  (코드 검증)

데이터 출처(우선순위): product_info.json 의 `bulk` 오버라이드 → product_info 본문 파생값 → CONFIG 기본값.
스토어 기본값(CONFIG)은 사용자 학습 예제(웨버네추럴스 L라이신) 기준:
  우체국택배(EPOST) · 수량별 6개당 15,000 · 재고 1000 · 신상품 · 과세상품 · 관부가세 포함.

사용:
  python3 scripts/build_bulk_excel.py <slug 또는 제품폴더경로>
  python3 scripts/build_bulk_excel.py webber_naturals_l_lysine_1000mg_140
  python3 scripts/build_bulk_excel.py --out /tmp/foo.xlsx output/new-item/<slug>

product_info.json 에 스토어 코드를 직접 박고 싶으면 `"bulk": {...}` 블록을 추가:
  "bulk": {
    "category_code": 50002615,      # 생략 시 category_proposed/fallback 경로로 자동 해석
    "origin_code": "0204006",       # 생략 시 origin.* 국가명으로 자동 해석
    "delivery_code": "EPOST",
    "shipping_type": "수량별", "shipping_fee": 15000, "shipping_qty": 6,
    "stock": 1000, "sale_price": 15800,
    "rep_image": "https://.../main.jpg",        # 대표이미지(필수) — 있으면 채움
    "return_fee": 5000, "exchange_fee": 5000    # 반품/교환배송비(조건부필수)
  }
"""
import argparse, glob, json, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GUIDE_DIR = os.path.join(ROOT, "guide", "일괄등록 guide")
NEW_ITEM = os.path.join(ROOT, "output", "new-item")

# 배송방법 칸 구분자는 일반 콤마가 아니라 U+201A(‚) — 템플릿 원본과 동일해야 함
SHIP_METHOD = "택배‚ 소포‚ 등기"

# 해외사업자 제약: 영양제류는 카테고리 선택이 '기타건강보조식품'만 가능(아미노산 등 세분류 못 고름).
# product_info 가 영양제 세분류(아미노산 등)를 제안해도 이 코드로 강제 + 경고 안 띄움.
SUPPLEMENT_FORCED_CODE = 50002615   # 식품 > 건강식품 > 영양제 > 기타건강보조식품

# 스토어 기본값 (사용자 학습 예제 기준; bulk 오버라이드/product_info 파생값이 우선)
CONFIG = {
    "product_state": "신상품",
    "tax_type": "과세상품",            # 부가세
    "import_tax": "관부가세 포함",       # 관부가세(K) 기본 = '포함' (2026-06-30) — 핀치마트는 국내발송 리셀러라 구매자 통관세 없음. 전 품목 포함.
    "unit_price_use": "N",            # 단위가격 사용여부
    "stock": 1000,
    "delivery_code": "EPOST",         # 우체국택배 (기본 — 사용자가 필요시 수동 변경)
    "origin_code": "0204006",         # 캐나다 (기본 — 전 상품 캐나다로 깔고 수동 보정)
    "shipping_type": "수량별",
    "shipping_fee": 15000,
    "shipping_qty": 6,                # 수량별부과-수량
    "shipping_pay": "선결제",          # 무조건 선결제 (사용자 지시 2026-06-07)
    "multi_origin": "N",
    # A/S·반품/교환 스토어 공통값 (항상 추가)
    "return_fee": 50000,
    "exchange_fee": 50000,
    "as_phone": "+16478010784",
    "as_guide": "수령시 파손된 제품에 한해 환불 가능합니다",
}

# 네이버 필드명(2행) → 데이터 dict 키. 컬럼 인덱스는 템플릿 2행에서 동적으로 찾는다.
FIELD = {
    "category_code": "카테고리코드",
    "title":        "상품명",
    "product_state":"상품상태",
    "sale_price":   "판매가",
    "unit_price_use":"단위가격 사용여부",
    "tax_type":     "부가세",
    "import_tax":   "관부가세",
    "stock":        "재고수량",
    "rep_image":    "대표이미지",
    "add_image":    "추가이미지",
    "detail":       "상세설명",
    "brand":        "브랜드",
    "maker":        "제조사",
    "origin_code":  "원산지코드",
    "importer":     "수입사",
    "multi_origin": "복수원산지여부",
    "ship_method":  "배송방법",
    "delivery_code":"택배사코드",
    "shipping_type":"배송비유형",
    "shipping_fee": "기본배송비",
    "shipping_pay": "배송비 결제방식",
    "shipping_qty": "수량별부과-수량",
    "return_fee":   "반품배송비",
    "exchange_fee": "교환배송비",
    "as_phone":     "A/S 전화번호",
    "as_guide":     "A/S 안내",
    # 상품정보제공고시 (기타 재화) — 개행 제거 후 매칭되므로 개행 없이 기재
    "gosi_name":    "상품정보제공고시품명",
    "gosi_model":   "상품정보제공고시모델명",
    "gosi_maker":   "상품정보제공고시제조자",
    # 옵션(단일상품+옵션 = 조합형/단독형)
    "option_type":  "옵션형태",
    "option_name":  "옵션명",
    "option_value": "옵션값",
    "option_price": "옵션가",
    "option_stock": "옵션 재고수량",
}


def latest(pattern):
    files = sorted(glob.glob(os.path.join(GUIDE_DIR, pattern)))
    if not files:
        sys.exit(f"[!] 파일 없음: {pattern} (in {GUIDE_DIR})")
    return files[-1]


def load_lookup_xls(path):
    import xlrd
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def category_candidates(pinfo):
    """product_info 스키마 변형(register·source-launch 여러 버전)에서 카테고리 경로 후보를 모은다.
    키 이름이 제각각(category_proposed·naver_category·category_naver·category_naver_confirmed…)이라
    '>' 가 든 경로형 문자열만 우선순위로 수집. note/flag 류(프로즈)는 제외."""
    PRIORITY = ["category_naver_confirmed", "category_naver_corrected", "category_proposed",
                "naver_category", "category_naver", "category_naver_alt", "category", "category_fallback"]
    seen = []

    def add(v):
        if isinstance(v, str) and ">" in v and v.strip() not in seen:
            seen.append(v.strip())

    for k in PRIORITY:
        add(pinfo.get(k))
    for k, v in pinfo.items():                    # 그 외 categor* 키(노트/플래그 제외)도 훑기
        kl = k.lower()
        if "categor" in kl and "note" not in kl and "flag" not in kl:
            add(v)
    return seen


def resolve_category(rows, *paths):
    """'식품 > 건강식품 > 영양제 > 아미노산' 경로 → 코드. 여러 경로를 순서대로 시도."""
    by_leaf = {}   # 세분류 leaf -> (code, full)
    by_full = {}
    for r in rows[1:]:
        code = str(r[0]).split(".")[0]
        chain = [str(x).strip() for x in r[1:] if str(x).strip()]
        if not chain:
            continue
        full = " > ".join(chain)
        by_full[full] = code
        by_leaf.setdefault(chain[-1], (code, full))
    for i, p in enumerate(paths):
        if not p:
            continue
        if p.strip() in by_full:
            return by_full[p.strip()], p.strip(), (i > 0)   # 정확 매칭 (i>0 이면 fallback)
    # leaf 만 매칭 (경로 일부 다를 때)
    for i, p in enumerate(paths):
        if not p:
            continue
        leaf = p.strip().split(">")[-1].strip()
        if leaf in by_leaf:
            code, full = by_leaf[leaf]
            return code, full, (i > 0)          # i>0 이면 fallback 사용
    return None, None, False


def resolve_shipping(pinfo):
    """product_info 의 shipping 정보 → (배송비유형, 기본배송비, 수량별부과-수량).

    스키마 변형 모두 흡수:
      - shipping dict: {per_unit_krw, per_N_units_krw, policy, absorbed_into_price}
      - shipping_krw / shipping(str): "2개당 15,000원 (개당 7,500원), 별도" 식 문자열
    찾지 못하면 (None, None, None) → 호출부에서 CONFIG 기본값으로 fallback.

    네이버 '수량별' 모델: 기본배송비가 '수량'개마다 부과.
      개당 N원        → 수량별, fee=N, qty=1
      M개당 N원       → 수량별, fee=N, qty=M
      무료/흡수        → 무료, fee=0, qty=None
    """
    import re

    def parse_str(s):
        if not s or not isinstance(s, str):
            return None
        if any(t in s for t in ("무료", "무배", "흡수")):
            return ("무료", 0, None)
        flat = s.replace(",", "")
        # 균일/유료 flat — "개수 상관없이"·"수량 무관"·"균일"·"주문당" N원 = 유료(수량 무관)
        if any(t in s for t in ("개수 상관없이", "개수상관없이", "수량 무관", "수량무관", "수량 상관없이", "균일", "주문당", "건당")):
            m = re.search(r"([\d]+)\s*원", flat)
            if m:
                return ("유료", int(m.group(1)), None)
        # "M개당 N원" 우선 (M>1 묶음). 콤마 제거 후 매칭.
        m = re.search(r"(\d+)\s*개당\s*([\d]+)\s*원?", flat)
        if m:
            return ("수량별", int(m.group(2)), int(m.group(1)))
        # "개당 N원" (수량 표기 없음 = 1개당)
        m = re.search(r"개당\s*([\d]+)\s*원?", flat)
        if m:
            return ("수량별", int(m.group(1)), 1)
        # "N원/개" (원/개 역순 표기 = 1개당)
        m = re.search(r"([\d]+)\s*원?\s*/\s*개", flat)
        if m:
            return ("수량별", int(m.group(1)), 1)
        # "수량당 N원" (묶음 없음 = 1개당 단순곱셈)
        m = re.search(r"수량당\s*([\d]+)\s*원?", flat)
        if m:
            return ("수량별", int(m.group(1)), 1)
        return None

    def from_dict(sh):
        if sh.get("absorbed_into_price") or sh.get("naver_free") is True:
            return ("무료", 0, None)
        # per_unit_krw / shipping_krw_per_unit 있으면 개당(수량별 qty=1) — 가장 명확한 신호
        per_unit = sh.get("per_unit_krw")
        if per_unit in (None, ""):
            per_unit = sh.get("shipping_krw_per_unit")
        # bundle_rule/policy/description/rule/note 문자열에 'M개당' 묶음이 명시돼 있으면 그쪽 우선
        r = parse_str(sh.get("policy") or sh.get("bundle_rule") or sh.get("description")
                      or sh.get("rule") or sh.get("note"))
        if r and r[2] not in (None, 1):
            return r
        if per_unit not in (None, ""):
            return ("수량별", int(per_unit), 1)
        # {bundle_fee_krw + per_bundle_qty} 묶음 스키마 — bundle_fee 가 per_bundle_qty 개마다 부과
        bf = sh.get("bundle_fee_krw")
        pq = sh.get("per_bundle_qty")
        if bf not in (None, "") and pq not in (None, ""):
            return ("수량별", int(bf), int(pq))
        # {krw + per_units} 묶음 스키마 (rule:"bundle") — krw 가 per_units 개마다 부과
        krw = sh.get("krw")
        per_units = sh.get("per_units")
        if krw not in (None, "") and per_units not in (None, ""):
            return ("수량별", int(krw), int(per_units))
        return r

    sh = pinfo.get("shipping")
    if isinstance(sh, dict):
        r = from_dict(sh)
        if r:
            return r
    elif isinstance(sh, str):
        r = parse_str(sh)
        if r:
            return r

    sp = pinfo.get("shipping_policy")
    if isinstance(sp, dict):
        r = from_dict(sp)
        if r:
            return r

    # 최상위 문자열 키들
    for key in ("shipping_krw", "shipping_policy", "shipping_fee_krw"):
        r = parse_str(pinfo.get(key))
        if r:
            return r
    return (None, None, None)


def clean_gosi_model(m, limit=50):
    """상품정보제공고시 모델명 정규화. 네이버 제약:
    ① 50자 제한 ② 등록불가 특수문자 \\ * ? " < > 금지.
    색상 접미(' - Color') 제거 → 금지문자 제거 → 공백 정리 →
    그래도 50자 넘치면 브랜드 접두(lululemon 등) 제거 → 최후 50자 컷."""
    if not m:
        return m
    base = str(m).split(" - ")[0].strip()          # 색상 접미 제거
    for ch in '\\*?"<>':                            # 네이버 등록불가 특수문자 제거
        base = base.replace(ch, "")
    while "  " in base:                             # 특수문자 제거로 생긴 이중 공백 정리
        base = base.replace("  ", " ")
    base = base.strip()
    if len(base) >= limit:                         # 50자 이상이면(에러 문구가 '50자보다 작게'=strictly<50) 브랜드 접두 제거
        for pre in ("lululemon ", "Lululemon ", "LULULEMON "):
            if base.startswith(pre):
                base = base[len(pre):].strip()
                break
    return base[:limit - 1].strip()                # 안전하게 49자 이하 보장


def build_data(pinfo, detail_html, cat_rows, deliv_rows):
    bulk = pinfo.get("bulk", {})
    warn = []

    def pick(key, *fallbacks):
        if key in bulk and bulk[key] not in (None, ""):
            return bulk[key]
        for fb in fallbacks:
            if fb not in (None, ""):
                return fb
        return CONFIG.get(key)

    pricing = pinfo.get("pricing") or {}
    if not isinstance(pricing, dict):
        pricing = {}

    # register 스키마(product_name_ko·sell_price_krw·brand_ko)와
    # source-launch 스키마(title_ko·sell_krw·brand) 둘 다 읽는다.
    brand_ko = pinfo.get("brand_ko") or pinfo.get("brand")
    d = {}
    d["title"] = pick("title", pinfo.get("product_name_ko"), pinfo.get("title_ko"), pinfo.get("group_name_ko"))
    # 판매가: pricing dict 내부 또는 최상위 키(스키마 변형: sell_price_krw·selling_price_krw·sell_krw)
    # 또는 calculation 중첩 블록(calculation.sell_price_krw / step4_sell_krw_raw) 모두 탐색
    calc = pinfo.get("calculation") if isinstance(pinfo.get("calculation"), dict) else {}
    d["sale_price"] = pick("sale_price",
                           pricing.get("sell_price_krw"), pricing.get("sell_krw"),
                           pinfo.get("sell_price_krw"), pinfo.get("selling_price_krw"), pinfo.get("sell_krw"),
                           calc.get("sell_price_krw"), calc.get("sell_krw"))
    d["brand"] = pick("brand", brand_ko)
    d["maker"] = pick("maker", brand_ko)
    d["importer"] = pick("importer", brand_ko)
    d["product_state"] = pick("product_state")
    d["tax_type"] = pick("tax_type")
    # import_tax(K 관부가세)는 카테고리 판정 후 설정 (아래 카테고리 블록 뒤)
    d["unit_price_use"] = pick("unit_price_use")
    d["stock"] = pick("stock")
    d["multi_origin"] = pick("multi_origin")
    d["ship_method"] = SHIP_METHOD
    d["delivery_code"] = pick("delivery_code")
    # 배송비: bulk 오버라이드 > product_info 파생값(resolve_shipping) > CONFIG 기본
    auto_st, auto_sf, auto_sq = resolve_shipping(pinfo)
    d["shipping_type"] = pick("shipping_type", auto_st)
    d["shipping_fee"] = pick("shipping_fee", auto_sf)
    d["shipping_pay"] = pick("shipping_pay")
    d["shipping_qty"] = pick("shipping_qty", auto_sq)
    # 수량별부과-수량은 '수량별' 유형에서만 유효. 유료(균일)·무료·조건부무료면 비움.
    if d["shipping_type"] != "수량별":
        d["shipping_qty"] = ""
    d["detail"] = detail_html
    d["rep_image"] = pick("rep_image")           # 보통 비어있음 → 사용자가 네이버에서 직접 업로드
    # 추가이미지(X) — 최대 9개, 줄바꿈(\n)으로 구분. bulk.add_images(list/str) 우선, 없으면 images.additional_image_urls
    add_imgs = bulk.get("add_images") or bulk.get("additional_images")
    if add_imgs is None:
        add_imgs = (pinfo.get("images") or {}).get("additional_image_urls")
    if isinstance(add_imgs, str):
        add_imgs = [add_imgs]
    d["add_image"] = "\n".join(str(u) for u in add_imgs[:9]) if add_imgs else None
    d["return_fee"] = pick("return_fee")         # 스토어 공통 50000
    d["exchange_fee"] = pick("exchange_fee")     # 스토어 공통 50000
    d["as_phone"] = pick("as_phone")             # A/S 전화번호 (스토어 공통)
    d["as_guide"] = pick("as_guide")             # A/S 안내 (스토어 공통)
    # 상품정보제공고시 — 품명=한글 상품명, 모델명=영문 공식 제품명, 제조자=브랜드
    d["gosi_name"] = pick("gosi_name", d.get("title"))
    d["gosi_model"] = clean_gosi_model(
        pick("gosi_model", pinfo.get("model_name"), pinfo.get("model_name_en"), pinfo.get("model_name_ko")))
    d["gosi_maker"] = pick("gosi_maker", brand_ko)

    # 옵션 — 단일상품에 옵션 여럿(options[]) 있으면 조합형으로 채움. 옵션값=color_ko(콤마구분),
    # 옵션가 전부 0(균일가), 옵션재고 전부 main 재고. (없으면 옵션칸 비움 = 단일무옵션)
    opts = pinfo.get("options")
    if isinstance(opts, list) and len(opts) >= 2:
        vals = [str(o.get("color_ko") or o.get("color_en") or o.get("option_label") or "").strip()
                for o in opts]
        vals = [v for v in vals if v]
        if vals:
            stock = str(d.get("stock") or CONFIG.get("stock") or 999)
            d["option_type"] = pick("option_type") or "조합형"
            d["option_name"] = pick("option_name") or "색상"
            d["option_value"] = ",".join(vals)
            d["option_price"] = ",".join(["0"] * len(vals))
            d["option_stock"] = ",".join([stock] * len(vals))

    # 카테고리코드 — bulk 우선 / 영양제는 해외사업자 제약상 기타건강보조식품 강제 / 그 외 경로 자동해석
    cat_candidates = category_candidates(pinfo)   # 스키마 변형 전반에서 경로 수집
    cat_paths = " ".join(cat_candidates)
    if bulk.get("category_code"):
        d["category_code"] = bulk["category_code"]
    elif "영양제" in cat_paths or "건강식품" in cat_paths:
        d["category_code"] = SUPPLEMENT_FORCED_CODE          # 기타건강보조식품 (해외사업자 제약상 유일 선택)
        warn.append("카테고리: 영양제 → 기타건강보조식품(50002615) 적용 [해외사업자 제약]. "
                    "더 맞는 카테고리 가능하면 등록화면에서 수동 변경")
    else:
        code, full, used_fb = resolve_category(cat_rows, *cat_candidates)
        if code:
            d["category_code"] = int(code) if str(code).isdigit() else code
            if used_fb:
                warn.append(f"카테고리: 1순위 경로 미발견 → fallback '{full}'({code}) 사용. 등록화면 확인")
        else:
            warn.append("카테고리코드 자동해석 실패 — bulk.category_code 직접 지정 필요")

    # K컬럼 관부가세 — 전 품목 '관부가세 포함'이 기본 (2026-06-30 사용자 지시: 핀치마트=국내발송 리셀러,
    # 구매자 통관세 없음). bulk.import_tax 명시가 있으면 그것을 최우선 존중(예외 케이스용).
    if bulk.get("import_tax"):
        d["import_tax"] = bulk["import_tax"]
    else:
        d["import_tax"] = CONFIG["import_tax"]      # 관부가세 포함

    # 원산지코드 — bulk 우선, 없으면 캐나다(0204006) 기본 고정 (전 상품 캐나다로 깔고 사용자 수동 보정)
    # (앞자리 0 보존 위해 문자열. 국가명 자동해석은 안 씀 — 사용자 지시로 캐나다 기본.)
    d["origin_code"] = str(bulk.get("origin_code") or CONFIG["origin_code"])

    # 택배사코드 검증
    valid_deliv = {str(r[0]) for r in deliv_rows[1:]}
    if d["delivery_code"] not in valid_deliv:
        warn.append(f"택배사코드 '{d['delivery_code']}' 가 코드 리스트에 없음 — 확인 필요")

    # 필수값 누락 경고 (구버전/미완성 product_info 방어)
    if not d.get("title"):
        warn.append("상품명(product_name_ko) 없음 — product_info 확인")
    if not d.get("sale_price"):
        warn.append("판매가(pricing.sell_price_krw) 없음 — product_info 확인")

    return d, warn


def write_excel(rows, out_path):
    """rows = 데이터 dict 의 리스트. 한 시트에 여러 SKU 를 3행부터 나열한다."""
    import openpyxl
    tpl = latest("ExcelSaveTemplate_*.xlsx")
    wb = openpyxl.load_workbook(tpl)
    ws = wb.worksheets[0]

    # 2행 필드명 → 컬럼 인덱스 매핑 (개행 제거 후 매칭)
    name2col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v:
            name2col[str(v).replace("\n", "").strip()] = c

    # 가이드 행(3~6) 삭제 → 데이터는 3행부터
    ws.delete_rows(3, 4)

    for i, d in enumerate(rows):
        r = 3 + i
        for key, val in d.items():
            if val in (None, ""):
                continue
            field = FIELD.get(key)
            col = name2col.get(field.replace("\n", "").strip()) if field else None
            if not col:
                continue
            ws.cell(row=r, column=col, value=val)
            if key == "origin_code":             # 앞자리 0 보존
                ws.cell(row=r, column=col).number_format = "@"
            if key == "add_image":               # 추가이미지 URL 여러 개 — 셀 내 줄바꿈 표시(엔터 구분)
                from openpyxl.styles import Alignment
                ws.cell(row=r, column=col).alignment = Alignment(wrapText=True, vertical="top")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def find_folder(arg):
    if os.path.isdir(arg):
        return arg
    cand = os.path.join(NEW_ITEM, arg)
    if os.path.isdir(cand):
        return cand
    sys.exit(f"[!] 제품 폴더를 찾을 수 없음: {arg}")


def one_row(target, lookups):
    """제품 폴더 1개 → (slug, 데이터dict, 경고). source-launch 배치/단건 공통."""
    folder = find_folder(target)
    slug = os.path.basename(os.path.normpath(folder))
    pinfo_path = os.path.join(folder, f"{slug}_product_info.json")
    detail_path = os.path.join(folder, f"{slug}_detail.html")
    if not os.path.exists(pinfo_path):
        sys.exit(f"[!] product_info 없음: {pinfo_path}")
    pinfo = json.load(open(pinfo_path, encoding="utf-8"))
    if not os.path.exists(detail_path):
        # 공통(공유) 상세 — product_info.detail_html 참조로 해석 (그룹/옵션 SKU 가 한 상세 공유)
        ref = pinfo.get("detail_html")
        resolved = None
        if ref:
            cands = [os.path.join(folder, ref),
                     os.path.join(NEW_ITEM, ref.replace("_detail.html", ""), ref)]
            resolved = next((c for c in cands if os.path.exists(c)), None)
            if not resolved:
                import glob
                hits = glob.glob(os.path.join(NEW_ITEM, "*", ref))
                resolved = hits[0] if hits else None
        if not resolved:
            sys.exit(f"[!] detail.html 없음: {detail_path} (detail_html ref={ref})")
        detail_path = resolved
    detail_html = open(detail_path, encoding="utf-8").read().strip()
    d, warn = build_data(pinfo, detail_html, *lookups)
    return slug, folder, d, warn


def main():
    ap = argparse.ArgumentParser(description="제품 산출물 → 네이버 일괄등록 엑셀 (단건/배치)")
    ap.add_argument("targets", nargs="+",
                    help="slug 또는 제품 폴더 경로. 여러 개 주면 한 엑셀에 다 담음(source-launch 배치)")
    ap.add_argument("--out", help="출력 xlsx 경로. 배치(2개+)면 기본 output/new-item/_batch/bulk_upload_<날짜>.xlsx")
    args = ap.parse_args()

    lookups = (load_lookup_xls(latest("category_*.xls")),
               load_lookup_xls(latest("delivery-companies_*.xls")))

    results = [one_row(t, lookups) for t in args.targets]
    batch = len(results) > 1

    if batch:
        if args.out:
            out = args.out
        else:
            from datetime import date
            out = os.path.join(NEW_ITEM, "_batch", f"bulk_upload_{date.today().isoformat()}.xlsx")
    else:
        out = args.out or os.path.join(results[0][1], f"{results[0][0]}_bulk_upload.xlsx")

    write_excel([r[2] for r in results], out)

    print(f"✅ 생성: {out}  ({len(results)}개 SKU, 데이터 {len(results)}행)")
    for slug, _folder, d, warn in results:
        print(f"\n· {slug}")
        print(f"    상품명: {d.get('title')}")
        print(f"    카테고리 {d.get('category_code')} · 판매가 {d.get('sale_price')} · 원산지 {d.get('origin_code')} · "
              f"배송 {d.get('delivery_code')}/{d.get('shipping_type')} {d.get('shipping_fee')}")
        for w in warn:
            print(f"    ⚠️ {w}")
        # 완구/어린이제품 → 어린이제품 KC 인증 (일괄엑셀 템플릿에 필드 없음, 업로드 후 네이버 UI 설정 필수)
        _cat = ""
        _pj = glob.glob(os.path.join(_folder, "*product_info.json"))
        if _pj:
            try:
                _pd = json.load(open(_pj[0], encoding="utf-8"))
                _cat = str(_pd.get("category_proposed") or _pd.get("naver_category") or "")
            except Exception:
                _cat = ""
        if any(k in _cat for k in ("완구", "유아", "어린이", "블록", "레고", "인형", "키즈", "출산/육아")):
            print("    🧸 완구/어린이제품 — 업로드 후 네이버에서 '어린이제품 인증 = 인증대상 아님' 설정 필수 "
                  "(일괄엑셀엔 인증 필드 없음 → 미설정 시 자동 판매중지)")
    missing_rep = [slug for slug, _f, d, _w in results if not d.get("rep_image")]
    if missing_rep:
        print("\n📌 사용자 직접 처리: 대표이미지(W) — 네이버에서 직접 업로드 "
              f"({', '.join(missing_rep)}). 반품/교환배송비·A/S 는 자동 채움")
    else:
        print("\n📌 대표이미지(W) 호스팅 URL 자동 입력 완료. 반품/교환배송비·A/S 자동 채움")


if __name__ == "__main__":
    main()
